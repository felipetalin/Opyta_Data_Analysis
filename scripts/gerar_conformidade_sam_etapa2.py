"""
Tabela de Conformidade Meio Fisico SAM Metais — Etapa 2 (validacao da migracao).

Layout WIDE (pivot):
    Header row 1: campanha (vazio nas colunas de parametro/unidade/VMPs;
                  nome da campanha mesclado sobre os pontos)
    Header row 2: nome_parametro, unidade_medida, <VMPs da matriz>, <ponto_1>, ...

Regras por matriz:
- Agua Superficial:  VMP_357_Cl1_Min, _Cl1_Max, _Cl2_Min, _Cl2_Max, _Cl3,
                     vmp_amonia_dinamico (CONAMA 357 — limite por pH).
- Agua Subterranea:  VMP_396_Consumo_Humano, _Dessedentacao_Animal,
                     _Irrigacao, _Recreacao.
- Sedimento:         VMP_454_N1, VMP_454_N2.

Marcacao de violacao:
- Cada celula de valor recebe fundo vermelho (FFC7CE) quando o valor
  numerico viola QUALQUER VMP aplicavel da linha.
- Para amonia, a violacao usa o pH medido no mesmo (campanha, ponto)
  e a tabela do CONAMA 357 (3.7 / 2.0 / 1.0 / 0.5 mg/L conforme pH).
- Sinal '<' (abaixo do LOQ) nunca configura violacao.

Fonte:
  Resultados_Meio_Fisico.xlsx (aba Resultados_Meio_Fisico)
  cadastro_parametros_opyta.xlsx (abas Aguas_Superficiais / Aguas_Subterraneas / Sedimento)

Saidas:
  Resultados/Meio_físico/<Superficial|Subterrânea|Sedimentos>/01_Conformidade_*.xlsx

Uso:
    python scripts/gerar_conformidade_sam_etapa2.py
"""

from __future__ import annotations

import os

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Caminhos
# ----------------------------------------------------------------------------
CLIENT_ROOT = Path(os.environ.get("OPYTA_MF_CLIENT_ROOT", r"G:/Meu Drive/Opyta/Clientes/Clientes/Clientes/Ferreira Rocha/SAM Metais/Produtos"))
SRC_RESULTADOS = CLIENT_ROOT / "Migração" / "Físico" / "Resultados_Meio_Fisico.xlsx"
SRC_CADASTRO = CLIENT_ROOT / "Migração" / "Físico" / "cadastro_parametros_opyta.xlsx"
OUT_ROOT = CLIENT_ROOT / "Resultados" / "Meio_físico"

PARAM_AMONIA = "Nitrogênio Amoniacal"
PARAM_PH = "pH In Situ"

# (col_no_cadastro_ou_None, label_no_excel, modo)
# modo: 'max' viola se valor > vmp ; 'min' viola se valor < vmp ; 'amonia' dinamico
MATRIZ_CFG = {
    "Água Superficial": {
        "subpasta": "Superficial",
        "aba_cad": "Aguas_Superficiais",
        "arquivo": "01_Conformidade_Agua_Superficial.xlsx",
        "vmps": [
            ("VMP_357_Cl1_Min", "vmp_357_cl1_min", "min"),
            ("VMP_357_Cl1_Max", "vmp_357_cl1_max", "max"),
            ("VMP_357_Cl2_Min", "vmp_357_cl2_min", "min"),
            ("VMP_357_Cl2_Max", "vmp_357_cl2_max", "max"),
            ("VMP_357_Cl3", "vmp_357_cl3", "max"),
            (None, "vmp_amonia_dinamico", "amonia"),
        ],
    },
    "Água Subterrânea": {
        "subpasta": "Subterrânea",
        "aba_cad": "Aguas_Subterraneas",
        "arquivo": "01_Conformidade_Agua_Subterranea.xlsx",
        "vmps": [
            ("VMP_396_Consumo_Humano", "vmp_396_consumo_humano", "max"),
            ("VMP_396_Dessedentacao_Animal", "vmp_396_dessedentacao_animal", "max"),
            ("VMP_396_Irrigacao", "vmp_396_irrigacao", "max"),
            ("VMP_396_Recreacao", "vmp_396_recreacao", "max"),
        ],
    },
    "Sedimento": {
        "subpasta": "Sedimentos",
        "aba_cad": "Sedimento",
        "arquivo": "01_Conformidade_Sedimento.xlsx",
        "vmps": [
            ("VMP_454_N1", "vmp_454_n1", "max"),
            ("VMP_454_N2", "vmp_454_n2", "max"),
        ],
    },
}

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FILL_LIGHT = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ----------------------------------------------------------------------------
# Parsers
# ----------------------------------------------------------------------------
def _parse_resultado(val) -> tuple[float | None, str]:
    """'< 0,5' -> (0.5, '<') ; '> 100' -> (100, '>') ; '12,3' -> (12.3, '')."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None, ""
    s = str(val).strip()
    sinal = ""
    if s.startswith("<"):
        sinal = "<"
        s = s[1:].strip()
    elif s.startswith(">"):
        sinal = ">"
        s = s[1:].strip()
    s = s.replace(",", ".").replace(" ", "")
    if not s:
        return None, sinal
    try:
        return float(s), sinal
    except ValueError:
        return None, sinal


def _parse_vmp(val) -> float | None:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip()
    if s in {"", "-", "—", "NA", "N/A"}:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _limite_amonia_por_ph(ph: float | None) -> float | None:
    """CONAMA 357 art.34 — Nitrogenio Amoniacal Total (mg/L N) por faixa de pH."""
    if ph is None:
        return None
    if ph <= 7.5:
        return 3.7
    if ph <= 8.0:
        return 2.0
    if ph <= 8.5:
        return 1.0
    return 0.5


def _viola(valor: float | None, sinal: str, vmp: float | None, modo: str) -> bool:
    if valor is None or vmp is None:
        return False
    if sinal == "<":
        return False
    if modo == "max":
        return valor > vmp
    if modo == "min":
        return valor < vmp
    return False


# ----------------------------------------------------------------------------
# Pipeline
# ----------------------------------------------------------------------------
def _ph_lookup(df: pd.DataFrame) -> dict[tuple, float | None]:
    """Mapa (campanha, ponto) -> pH medido (ou None)."""
    sub = df[df["Parametro"] == PARAM_PH]
    out: dict[tuple, float | None] = {}
    for _, row in sub.iterrows():
        v, sinal = _parse_resultado(row["Resultado"])
        out[(row["Campanha"], row["Ponto"])] = v if sinal != "<" else None
    return out


def gerar_para_matriz(df_res: pd.DataFrame, matriz: str, cfg: dict) -> tuple[Path, int, int]:
    df = df_res[df_res["Matriz"] == matriz].copy()
    df_cad = pd.read_excel(SRC_CADASTRO, sheet_name=cfg["aba_cad"])
    cad_idx = df_cad.set_index("Parametro")

    ph_map = _ph_lookup(df) if matriz == "Água Superficial" else {}

    # Mapas VMP por parametro
    vmp_maps: dict[str, dict] = {}
    for col_cad, label, _modo in cfg["vmps"]:
        if col_cad is None or col_cad not in df_cad.columns:
            continue
        vmp_maps[label] = {p: _parse_vmp(v) for p, v in cad_idx[col_cad].items()}

    # Parse resultado
    parsed = df["Resultado"].map(_parse_resultado)
    df["_valor"] = parsed.map(lambda t: t[0])
    df["_sinal"] = parsed.map(lambda t: t[1])

    parametros = sorted(df["Parametro"].unique())
    campanhas = sorted(df["Campanha"].unique())
    pontos = sorted(df["Ponto"].unique())

    # Unidade por parametro
    unidades: dict[str, str] = {}
    for p in parametros:
        u = None
        if "Unidade_Medida" in df_cad.columns and p in cad_idx.index:
            u = cad_idx.loc[p, "Unidade_Medida"]
            if isinstance(u, pd.Series):
                u = u.iloc[0]
        if u is None or (isinstance(u, float) and np.isnan(u)):
            ds = df.loc[df["Parametro"] == p, "Unidade_Medida"]
            u = ds.iloc[0] if len(ds) else ""
        unidades[p] = u

    # Lookups por (parametro, campanha, ponto)
    res_str: dict[tuple, str] = {}
    res_num: dict[tuple, tuple[float | None, str]] = {}
    for _, row in df.iterrows():
        k = (row["Parametro"], row["Campanha"], row["Ponto"])
        res_str[k] = row["Resultado"] if pd.notna(row["Resultado"]) else ""
        res_num[k] = (row["_valor"], row["_sinal"])

    # ----------------------------------------------------------------
    # Monta header e linhas
    # ----------------------------------------------------------------
    vmp_labels = [lbl for _, lbl, _ in cfg["vmps"]]
    header2 = ["nome_parametro", "unidade_medida"] + vmp_labels
    for _c in campanhas:
        header2 += list(pontos)

    header1: list = [""] * (2 + len(vmp_labels))
    for c in campanhas:
        header1 += [c] * len(pontos)

    rows: list[list] = []
    for p in parametros:
        linha: list = [p, unidades.get(p, "")]
        for _col_cad, label, modo in cfg["vmps"]:
            if modo == "amonia":
                linha.append(
                    "3.7 / 2.0 / 1.0 / 0.5 mg/L (pH ≤7.5 / ≤8 / ≤8.5 / >8.5)"
                    if p == PARAM_AMONIA else ""
                )
            else:
                v = vmp_maps.get(label, {}).get(p)
                linha.append(v if v is not None else "")
        for c in campanhas:
            for pt in pontos:
                linha.append(res_str.get((p, c, pt), ""))
        rows.append(linha)

    # ----------------------------------------------------------------
    # Escreve Excel
    # ----------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Conformidade"
    ws.append(header1)
    ws.append(header2)
    for r in rows:
        ws.append(r)

    # Mescla campanha sobre seus pontos
    start_col = 2 + len(vmp_labels) + 1
    for i, _c in enumerate(campanhas):
        c1 = start_col + i * len(pontos)
        c2 = c1 + len(pontos) - 1
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)

    # Estilo do header
    for col in range(1, len(header2) + 1):
        c1 = ws.cell(row=1, column=col)
        c2 = ws.cell(row=2, column=col)
        c1.fill = HEADER_FILL
        c1.font = HEADER_FONT
        c1.alignment = CENTER
        c2.fill = HEADER_FILL_LIGHT
        c2.font = HEADER_FONT
        c2.alignment = CENTER

    # Pinta celulas em violacao
    viol_cells = 0
    viol_params: set[str] = set()
    data_start_col = 2 + len(vmp_labels) + 1

    for ridx, p in enumerate(parametros, start=3):
        ativos: list[tuple[str, float | None, str]] = []
        for _col_cad, label, modo in cfg["vmps"]:
            if modo == "amonia":
                ativos.append((label, None, "amonia"))
            else:
                v = vmp_maps.get(label, {}).get(p)
                if v is not None:
                    ativos.append((label, v, modo))

        col = data_start_col
        for c in campanhas:
            for pt in pontos:
                valor, sinal = res_num.get((p, c, pt), (None, ""))
                violou = False
                for _lbl, vmp_val, modo in ativos:
                    if modo == "amonia":
                        if p != PARAM_AMONIA:
                            continue
                        lim = _limite_amonia_por_ph(ph_map.get((c, pt)))
                        if _viola(valor, sinal, lim, "max"):
                            violou = True
                            break
                    elif _viola(valor, sinal, vmp_val, modo):
                        violou = True
                        break
                if violou:
                    ws.cell(row=ridx, column=col).fill = RED_FILL
                    viol_cells += 1
                    viol_params.add(p)
                col += 1

    # Larguras
    ws.column_dimensions[get_column_letter(1)].width = 42
    ws.column_dimensions[get_column_letter(2)].width = 14
    for i in range(3, 2 + len(vmp_labels) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    for i in range(data_start_col, data_start_col + len(campanhas) * len(pontos)):
        ws.column_dimensions[get_column_letter(i)].width = 11

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = ws.cell(row=3, column=3)

    out_dir = OUT_ROOT / cfg["subpasta"]
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / cfg["arquivo"]
    try:
        wb.save(out_path)
    except PermissionError:
        alt = out_path.with_name(out_path.stem + "_NEW" + out_path.suffix)
        wb.save(alt)
        print(f"  ! arquivo original bloqueado; salvo em: {alt.name}")
        out_path = alt

    return out_path, viol_cells, len(viol_params)


def main() -> None:
    print(f"[etapa-2] Lendo {SRC_RESULTADOS.name} ...")
    df_res = pd.read_excel(SRC_RESULTADOS, sheet_name="Resultados_Meio_Fisico")
    print(f"  -> {len(df_res)} linhas | matrizes: {sorted(df_res['Matriz'].unique())}")
    print()

    for matriz, cfg in MATRIZ_CFG.items():
        print(f"[etapa-2] Processando {matriz} (layout WIDE) ...")
        out_path, viol_cells, viol_params = gerar_para_matriz(df_res, matriz, cfg)
        print(f"  -> {out_path}")
        print(f"     celulas em violacao: {viol_cells}  |  parametros afetados: {viol_params}")
        print()

    print("[etapa-2] OK")


if __name__ == "__main__":
    main()
